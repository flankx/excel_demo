#!/usr/bin/python
# coding=utf-8

import sys
import xlrd
import xlwt
import openpyxl
import hashlib
from xlutils.copy import copy


'''
check args from command
    param1: filename，文件名，如 /tmp/test.txt or /tmp/tetx.xls
'''
if len(sys.argv) == 2:
    filename = sys.argv[1]
else:
    print 'usage: python ' + __file__ + ' param1:filename '
    sys.exit()


# 处理 .xls 文件
def r_w_excel_xls(path):

    # 打开文件，获取excel文件的workbook（工作簿）对象
    workbook = xlrd.open_workbook(path)

    '''
    Workbook类初始化时有encoding和style_compression参数
    encoding:设置字符编码，一般要这样设置：w = Workbook(encoding='utf-8')，就可以在excel中输出中文了。默认是ascii。
    style_compression:表示是否压缩，不常用。
    '''
    # 创建一个Workbook对象，相当于创建了一个Excel文件
    book = xlwt.Workbook(encoding="utf-8", style_compression=2)

    '''对workbook对象进行操作'''
    # 获取所有sheet的名字
    names = workbook.sheet_names()
    print(names)
    # 通过sheet索引获得sheet对象
    worksheet = workbook.sheet_by_index(0)
    print(worksheet)

    '''对sheet对象进行操作'''
    name = worksheet.name  # 获取表的姓名
    nrows = worksheet.nrows  # 获取该表总行数
    ncols = worksheet.ncols  # 获取该表总列数
    print(name, nrows, ncols)

    sheet = book.add_sheet(name, cell_overwrite_ok=True)

    for i in range(nrows):
        print 'start write rows : {}'.format(i)
        row = worksheet.row_values(i)
        if i == 0:
            row.append('encrypt_phone')
        elif row[1]::
            row.append(hashlib.md5(row[1]))
        print i, row
        for j in range(ncols+1):
            try:
                # 选择日期的样式
                style1 = xlwt.easyxf(num_format_str='YYYY/MM/DD')
                style2 = xlwt.easyxf(num_format_str='hh:mm:ss')
                if j == 2:
                    sheet.write(i, j, row[j], style1)
                elif j == 3:
                    sheet.write(i, j, row[j], style2)
                else:
                    sheet.write(i, j, row[j])
            except Exception, e:
                print e
                continue

    book.save('/tmp/encrypt_result.xls')


# 处理 .xlsx 文件
def r_w_excel_xlsx(path):
    # 加载已经存在的工作簿 (workbook)
    workbook = openpyxl.load_workbook(path)
    # 列出所有的工作表 (sheet)
    sheets = workbook.sheetnames
    print sheets

    # 获取该表相应的行数和列数
    sheet = workbook[sheets[0]]
    print(sheet.max_row, sheet.max_column)

    # 在第一列之前插入一列
    sheet.insert_cols(0)

    for index, row in enumerate(sheet):
        print index
        if index == 0:
            row[0].value = 'phone_decrypt'
        elif row[2].value:
            row[0].value = hashlib.md5(row[2].value)
    # 保存至原文件
    workbook.save(filename)


if __name__ == '__main__':

    if filename.endswith('.xls'):
        r_w_excel_xls(filename)
    elif filename.endswith('.xlsx'):
        r_w_excel_xlsx(filename)
    print 'success'


"""
EXCEL 记录示例
lid	                data	                            date	    time
1200012813030401	3A884FA5D75A2EB61E8785624DCE3306	2018/12/19	20:54:14

"""
